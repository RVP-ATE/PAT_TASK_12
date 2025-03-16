# Import necessary modules and libraries
from datetime import datetime  # For handling date and time
import pytest  # For the testing framework (pytest)
from selenium import webdriver  # For automating browser interactions with Selenium
from openpyxl import load_workbook  # For reading and writing Excel files
from LoginPage import LoginPage  # Import the LoginPage class (which is assumed to be defined elsewhere)


# Fixture for setting up and tearing down the WebDriver instance
@pytest.fixture
def setup():
    # Initialize the Chrome WebDriver and navigate to the login page
    driver = webdriver.Chrome()
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")  # URL of the login page
    driver.maximize_window()  # Maximize the browser window for better visibility
    yield driver  # Yield the driver for use in the test functions
    driver.quit()  # Quit the browser after the test has finished


# Function to read test data from an Excel file
def read_test_data(file_path):
    wb = load_workbook(file_path)  # Open the Excel workbook
    ws = wb.active  # Access the active worksheet
    data = []  # Initialize an empty list to store test data

    # Loop through all rows starting from row 2 (skipping the header row) and retrieve data
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)  # Add each row of data to the list

    return data, wb, ws  # Return the data, workbook, and worksheet


# Function to write the test result back to the Excel file
def write_test_result(ws, row, result):
    # Write the test result to column 7 (Test Result), time to column 5 (Time), and date to column 4 (Date)
    ws.cell(row=row, column=7, value=result)  # Write the test result (Passed/Failed)
    ws.cell(row=row, column=5, value=str(datetime.now().strftime('%H:%M:%S')))  # Current time in HH:MM:SS format
    ws.cell(row=row, column=4, value=str(datetime.now().strftime('%Y-%m-%d')))  # Current date in YYYY-MM-DD format


# The actual test function
def test_login(setup):
    file_path = 'login_data.xlsx'  # Path to the test data file
    data, wb, ws = read_test_data(file_path)  # Read the test data from the Excel file

    # Loop through the test data, starting from row 2 in the Excel sheet
    for index, (test_id, username, password, _, _, _, _) in enumerate(data, start=2):
        login_page = LoginPage(setup)  # Initialize the LoginPage class with the driver setup

        # Perform login using the username and password from the data
        login_page.login(username, password)

        # Verify if the login was successful
        if login_page.is_login_successful():
            write_test_result(ws, index, "Test Passed")  # Write 'Test Passed' if login was successful
        else:
            write_test_result(ws, index, "Test Failed")  # Write 'Test Failed' if login was unsuccessful

    wb.save(file_path)  # Save the modified Excel file with the updated test results
